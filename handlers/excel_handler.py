import openpyxl
import time
from openpyxl.utils.exceptions import IllegalCharacterError
import pandas as pd


class ExcelHandler:
    def __init__(self, file_path, max_retries=3):
        self.file_path = file_path
        self.max_retries = max_retries
        self.wb, self.ws = self.load_excel()

    def load_excel(self):
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            return wb, ws
        except Exception as e:
            print(f"Failed to load Excel file: {str(e)}")
            return None, None

    def save_to_excel(self):
        for attempt in range(self.max_retries):
            try:
                self.wb.save(self.file_path)
                print(f"Excel file saved successfully.")
                return True
            except PermissionError:
                print(
                    f"PermissionError: Attempt {attempt + 1} of {self.max_retries}. Retrying in 5 seconds..."
                )
                time.sleep(5)
        print(f"Failed to save Excel file after {self.max_retries} attempts.")
        return False

    def update_cell(self, row, column, value):
        """
        指定されたセルの内容を1箇所だけ更新する。

        :param row: 更新対象の行番号
        :param column: 更新対象の列番号
        :param value: セルに設定する値
        :return: 更新が成功したかどうか
        """
        try:
            self.ws.cell(row=row, column=column, value=value)
            return True
        except IllegalCharacterError:
            print(
                f"IllegalCharacterError: Unable to save content due to invalid characters."
            )
            return False
        except Exception as e:
            print(f"An error occurred while updating the cell: {str(e)}")
            return False

    def update_cells(self, target_index, target_indices, values, row_flag=True):
        """
        指定された行または列の複数のセルを更新する。

        :param target_index: 更新対象の行番号または列番号
        :param target_indices: 更新対象の列番号または行番号のリスト
        :param values: 対応するセルに設定する値のリスト
        :param row_flag: Trueなら行を指定し、Falseなら列を指定
        :return: 更新が成功したかどうか
        """
        try:
            if row_flag:
                # 行を指定して、複数の列を更新
                for column, value in zip(target_indices, values):
                    self.ws.cell(row=target_index, column=column, value=value)
            else:
                # 列を指定して、複数の行を更新
                for row, value in zip(target_indices, values):
                    self.ws.cell(row=row, column=target_index, value=value)
            return True
        except IllegalCharacterError:
            print(
                f"IllegalCharacterError: Unable to save content due to invalid characters."
            )
            return False
        except Exception as e:
            print(f"An error occurred while updating the cells: {str(e)}")
            return False

    def find_matching_index(self, index, search_string, is_row_flag):
        """
        行または列を指定して検索し、完全に一致するインデックスを返す。

        :param index: 行番号または列番号 (1から始まる)
        :param search_string: 検索する文字列
        :param is_row_flag: Trueなら行で検索、Falseなら列で検索
        :return: 完全に一致する列番号または行番号、一致しない場合はNone
        """
        if is_row_flag:
            for col in range(1, self.ws.max_column + 1):
                cell_value = self.ws.cell(row=index, column=col).value
                if (
                    cell_value is not None
                    and str(cell_value).strip() == str(search_string).strip()
                ):
                    return col
        else:
            for row in range(1, self.ws.max_row + 1):
                cell_value = self.ws.cell(row=row, column=index).value
                if (
                    cell_value is not None
                    and str(cell_value).strip() == str(search_string).strip()
                ):
                    return row
        return None

    def is_cell_empty_or_match(self, row, column, expected_value=None):
        """
        セルが空か、オプションの値と一致するかを確認する。

        :param row: チェック対象の行番号
        :param column: チェック対象の列番号
        :param expected_value: 確認する値 (オプション)
        :return: セルが空か、または expected_value と一致する場合は True、それ以外は False
        """
        cell_value = self.ws.cell(row=row, column=column).value

        # pd.isna のチェックを行う
        if pd.isna(cell_value):
            return True

        # セルが空かどうかのチェック
        if cell_value is None or cell_value == "":
            return True

        # expected_value が指定されている場合のチェック
        if expected_value is not None:
            return str(cell_value) == str(expected_value)

        return False
