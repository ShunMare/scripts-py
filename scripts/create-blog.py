import os
import time
import pyautogui
from dotenv import load_dotenv
import pandas as pd
import pyperclip
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError
import pygetwindow as gw

# .envファイルから環境変数を読み込む
load_dotenv()

# Excelファイルのパスを環境変数から取得
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH")
WAIT_TIME_AFTER_PROMPT_LONG = int(os.getenv("WAIT_TIME_AFTER_PROMPT_LONG"))
WAIT_TIME_AFTER_PROMPT_SHORT = int(os.getenv("WAIT_TIME_AFTER_PROMPT_SHORT"))
WAIT_TIME_AFTER_RELOAD = int(os.getenv("WAIT_TIME_AFTER_RELOAD"))
WAIT_TIME_AFTER_SWITCH = int(os.getenv("WAIT_TIME_AFTER_SWITCH"))
print(f"Excel file path: {EXCEL_FILE_PATH}")

# Excelファイルを読み込む
df = pd.read_excel(EXCEL_FILE_PATH)

# 列名を小文字に変換
df.columns = df.columns.str.lower()

print("Available columns:", df.columns)
print("\nFirst few rows of the DataFrame:")
print(df.head())


def activate_edge():
    edge_windows = gw.getWindowsWithTitle("Edge")
    if edge_windows:
        edge_window = edge_windows[0]
        if not edge_window.isActive:
            edge_window.activate()
        print("Edge window activated")
        time.sleep(WAIT_TIME_AFTER_SWITCH)
    else:
        print("No Edge window found")


def save_to_excel(wb, path, retries=3):
    for attempt in range(retries):
        try:
            wb.save(path)
            print(f"Excel file saved successfully.")
            return True
        except PermissionError:
            print(
                f"PermissionError: Attempt {attempt + 1} of {retries}. Retrying in 5 seconds..."
            )
            time.sleep(5)
    print(f"Failed to save Excel file after {retries} attempts.")
    return False

def move_to_generate_button():
    for _ in range(3):
        pyautogui.hotkey("shift", "tab")
        time.sleep(0.5)
    print("move to generate button")

def move_to_copy_button():
    for _ in range(7):
        pyautogui.hotkey("shift", "tab")
        time.sleep(0.5)
    print("move to copy button")

def generate_and_process_prompts(group, start_row):
    first_row = group.iloc[0]

    if (
        "flag" not in first_row
        or pd.isna(first_row["flag"])
        or first_row["flag"] != 1.0
    ):
        print(f"Skipping group: No valid flag found. First row: {first_row}")
        return

    # Edgeウィンドウをアクティブにする (グループ処理開始時)
    activate_edge()

    theme = first_row.get("theme", "")
    heading = first_row.get("heading", "") if pd.notna(first_row.get("heading")) else ""

    evidences = [
        row["evidence"]
        for _, row in group.iterrows()
        if pd.notna(row["evidence"]) and row["evidence"] != ""
    ]

    if len(evidences) == 0:
        print(f"Skipping group: No evidences found. First row: {first_row}")
        return

    # 最初のプロンプトを生成
    initial_prompt = f"{theme}について書きたい。これまでの内容は入れないで、また一から書いて"
    if heading:
        initial_prompt += f"、見出しは{heading}です"
    initial_prompt += "。以下の指示に従って作成してください：\n"
    initial_prompt += "• できるだけわかりやすく、長く、ブログ形式で書いてください。\n"
    initial_prompt += "• 下記の内容以外から情報を出さないでください（ハルシネーション防止のため）。\n"
    initial_prompt += "• 見出しの番号は除いてください。\n\n"
    initial_prompt += "• タイトルは記載しないで、見出しはh2（##）、h3（###）、h4（####）で構成してください。\n\n"
    initial_prompt += "• 最後の文章に記事に合う絵文字を３つ付けてください。\n\n"
    initial_prompt += "• 下記の内容を参照してください。\n"
    initial_prompt += "参照内容：\n" + evidences[0]
    # プロンプトをクリップボードにコピーして、Edgeに貼り付け
    pyperclip.copy(initial_prompt)
    print("Initial prompt copied to clipboard.")

    # Edgeウィンドウをアクティブにする
    activate_edge()

    # Ctrl+R でページをリロード
    pyautogui.hotkey("ctrl", "r")
    time.sleep(WAIT_TIME_AFTER_RELOAD)

    # プロンプトを貼り付けて実行
    pyautogui.hotkey("ctrl", "v")
    time.sleep(1)
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.press("enter")
    time.sleep(WAIT_TIME_AFTER_PROMPT_LONG)
    move_to_generate_button()
    pyautogui.hotkey("space")
    time.sleep(WAIT_TIME_AFTER_PROMPT_SHORT)

    # 残りのエビデンスを処理
    for evidence in evidences[1:]:
        # Ctrl+R でページをリロード
        pyautogui.hotkey("ctrl", "r")
        time.sleep(WAIT_TIME_AFTER_RELOAD)
        additional_prompt = f"下記の内容を上記に加えて。省略せずに全部書いて。\n{evidence}"
        pyperclip.copy(additional_prompt)
        print("Additional prompt copied to clipboard.")

        # Edgeウィンドウをアクティブにする
        activate_edge()

        # プロンプトを貼り付けて実行
        pyautogui.hotkey("ctrl", "v")
        time.sleep(1)
        pyautogui.press("tab")
        time.sleep(0.5)
        pyautogui.press("enter")
        time.sleep(WAIT_TIME_AFTER_PROMPT_LONG)
        # 生成を続ける２回繰り返す
        for _ in range(2):
            move_to_generate_button()
            pyautogui.hotkey("space")
            time.sleep(WAIT_TIME_AFTER_PROMPT_LONG)

    # Ctrl+R でページをリロード
    pyautogui.hotkey("ctrl", "r")
    time.sleep(WAIT_TIME_AFTER_RELOAD)

    move_to_copy_button()
    pyautogui.hotkey("space")

    # "md"列に貼り付け
    generated_content = pyperclip.paste()
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # 'md'列が存在するか確認し、なければ作成
        md_column = None
        title_column = None
        description_column = None
        keywords_column = None
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value:
                if col[0].value.lower() == "md":
                    md_column = col[0].column
                elif col[0].value.lower() == "title":
                    title_column = col[0].column
                elif col[0].value.lower() == "description":
                    description_column = col[0].column
                elif col[0].value.lower() == "link":
                    link_column = col[0].column
                elif col[0].value.lower() == "keywords":
                    keywords_column = col[0].column

        if md_column is None:
            md_column = ws.max_column + 1
            ws.cell(row=1, column=md_column, value="md")
            print("Created 'md' column")

        # flagが1.0の行を見つける（これは常にグループの最初の行になるはず）
        flag_row = start_row + 2  # Excelの行番号は1から始まるため、2を加える

        ws.cell(row=flag_row, column=md_column, value=generated_content)

        def generate_and_paste_metadata(prompt, column):
            if column is None:
                print(f"Column for {prompt} not found. Skipping.")
                return

            pyperclip.copy(prompt)
            print(f"{prompt} copied to clipboard.")

            activate_edge()
            pyautogui.hotkey("ctrl", "r")
            time.sleep(WAIT_TIME_AFTER_RELOAD)
            pyautogui.hotkey("ctrl", "v")
            time.sleep(1)
            pyautogui.press("tab")
            time.sleep(0.5)
            pyautogui.press("enter")
            time.sleep(WAIT_TIME_AFTER_PROMPT_SHORT)
            pyautogui.hotkey("ctrl", "r")
            time.sleep(WAIT_TIME_AFTER_RELOAD)

            move_to_copy_button()
            pyautogui.hotkey("space")

            metadata_content = pyperclip.paste()
            ws.cell(row=flag_row, column=column, value=metadata_content)
            print(
                f"Generated {prompt} pasted to Excel (row {flag_row}, column {column})"
            )

        generate_and_paste_metadata(f"上記のタイトルを考えて。下記のキーワードを入れて\n---\n{theme}", title_column)
        generate_and_paste_metadata("パーマリンクを考えて", link_column)
        generate_and_paste_metadata(
            "上記のメタディスクリプションを長めに考えて", description_column
        )
        generate_and_paste_metadata(
            "もう少し短くして", description_column
        )
        generate_and_paste_metadata("上記のメタキーワードを考えてください。,で区切って書いてください。", keywords_column)

        # 画像生成のプロンプト
        image_prompt = "記事に合う横長の画像を作ってください。"
        pyperclip.copy(image_prompt)
        print(f"Image generation prompt: {image_prompt}")

        activate_edge()
        pyautogui.hotkey("ctrl", "r")
        time.sleep(WAIT_TIME_AFTER_RELOAD)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(1)
        pyautogui.press("tab")
        time.sleep(0.5)
        pyautogui.press("enter")
        time.sleep(WAIT_TIME_AFTER_PROMPT_SHORT)

        print("Image generation prompt sent. Image should be generated but not saved.")

        if save_to_excel(wb, EXCEL_FILE_PATH):
            print(f"All content pasted to Excel (row {flag_row})")
        else:
            print(f"Failed to save content to Excel (row {flag_row})")
    except IllegalCharacterError:
        print(
            f"IllegalCharacterError: Unable to save content due to invalid characters. Skipping this entry."
        )
    except Exception as e:
        print(f"An error occurred while saving to Excel: {str(e)}")


# 10行ごとにグループ化して処理
grouped = df.groupby(df.index // 10)

for i, (_, group) in enumerate(grouped):
    start_row = i * 10
    print(f"\nProcessing group starting at row {start_row}")
    generate_and_process_prompts(group, start_row)

print("\nAll prompts have been processed and results saved to Excel.")
