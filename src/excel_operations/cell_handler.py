from openpyxl.utils.exceptions import IllegalCharacterError
import pandas as pd
from src.log_operations.log_handlers import CustomLogger

logger = CustomLogger(__name__)


class ExcelCellHandler:
    def __init__(self):
        """
        ExcelCellHandlerの初期化
        """
        self.logger = CustomLogger(__name__)
        self.worksheet = None

    def set_worksheet(self, worksheet):
        """
        ワークシートを設定します。
        :param worksheet: 対象のワークシート
        """
        self.worksheet = worksheet
        self.logger.debug("Worksheet has been set successfully.")

    def update_cell(self, row, column, value):
        """
        指定されたセルに値を更新します。
        :param row: 行番号
        :param column: 列番号
        :param value: 書き込む値
        :return: True: 成功、False: エラー発生
        """
        try:
            self.worksheet.cell(row=row, column=column, value=value)
            logger.debug(f"Successfully updated cell at row {row}, column {column}")
            return True
        except IllegalCharacterError:
            logger.error(
                "IllegalCharacterError: Unable to save content due to invalid characters."
            )
            return False
        except Exception as e:
            logger.error(f"An error occurred while updating the cell: {str(e)}")
            return False

    def is_cell_empty_or_match(self, row, column, expected_value=None):
        """
        指定されたセルが空か、指定された値と一致するかを確認します。
        :param row: 行番号
        :param column: 列番号
        :param expected_value: チェックしたい期待値（任意）
        :return: True: 空または一致、False: 不一致
        """
        cell_value = self.worksheet.cell(row=row, column=column).value
        if pd.isna(cell_value) or cell_value is None or cell_value == "":
            logger.debug(f"Cell at row {row}, column {column} is empty")
            return True
        if expected_value is not None:
            result = str(cell_value) == str(expected_value)
            logger.debug(
                f"Cell at row {row}, column {column} {'matches' if result else 'does not match'} expected value"
            )
            return result
        logger.debug(f"Cell at row {row}, column {column} is not empty")
        return False

    def iterate_column_values(self, column: int, start_row: int = 1):
        """
        指定された列の指定された行から最終行までを繰り返し、行番号と値を返すジェネレータ。
        :param column: 列番号 (例: 3はC列)
        :param start_row: 開始行 (デフォルトは1行目)
        :yield: (行番号, セルの値) のタプル
        """
        logger.debug(f"Iterating column {column} values from row {start_row}")
        for row in range(start_row, self.worksheet.max_row + 1):
            value = self.worksheet.cell(row=row, column=column).value
            logger.debug(f"Row {row}, Column {column}: {value}")
            yield row, value

    def get_column_values_to_last_row(self, column: int, start_row: int = 1):
        """
        指定された列の指定された行から最終行までのセルの値を配列で返す。
        :param worksheet: 対象のワークシート
        :param column: 列番号 (例: 3はC列)
        :param start_row: 開始行 (デフォルトは1行目)
        :return: 指定された列の開始行から最終行までの値を含むリスト
        """
        values = []
        for row in range(start_row, self.worksheet.max_row + 1):
            value = self.worksheet.cell(row=row, column=column).value
            values.append(value)
        logger.debug(
            f"Retrieved {len(values)} values from column {column}, starting at row {start_row}"
        )
        return values

    def get_cell_value(self, row: int, column: int):
        """
        指定されたセルの値を取得します。
        :param worksheet: 対象のワークシート
        :param row: 行番号
        :param column: 列番号
        :return: セルの値
        """
        try:
            value = self.worksheet.cell(row=row, column=column).value
            logger.debug(f"Retrieved value from row {row}, column {column}: {value}")
            return value
        except Exception as e:
            logger.error(
                f"An error occurred while getting the cell value at row {row}, column {column}: {str(e)}"
            )
            return None

    def get_cell_value_by_column_letter(self, row: int, column_letter: str):
        """
        列文字（A, B, C...）を使用して指定されたセルの値を取得します。
        :param worksheet: 対象のワークシート
        :param row: 行番号
        :param column_letter: 列文字（例: 'A', 'B', 'C'...）
        :return: セルの値
        """
        try:
            value = self.worksheet[f"{column_letter}{row}"].value
            logger.debug(
                f"Retrieved value from row {row}, column {column_letter}: {value}"
            )
            return value
        except Exception as e:
            logger.error(
                f"An error occurred while getting the cell value at row {row}, column {column_letter}: {str(e)}"
            )
            return None

    def get_range_values(self, start_row: int, column: int, num_rows: int):
        """
        指定された行の指定された列から、指定された行数分のデータを配列で返します。
        :param start_row: 開始行
        :param column: 列番号
        :param num_rows: 取得する行数
        :return: セルの値のリスト
        """
        values = []
        for row in range(start_row, start_row + num_rows):
            value = self.worksheet.cell(row=row, column=column).value
            values.append(value)
        logger.debug(
            f"Retrieved {len(values)} values from column {column}, starting at row {start_row}, for {num_rows} rows"
        )
        return values

    def get_last_row_of_column(self, column: int) -> int:
        """
        指定された列の最終行（データが存在する最後の行）の行番号を返します。
        :param column: 列番号 (例: 3はC列)
        :return: 最終行の行番号
        """
        last_row = self.worksheet.max_row
        while last_row > 0:
            if self.worksheet.cell(row=last_row, column=column).value is not None:
                logger.debug(f"Last row with data in column {column}: {last_row}")
                return last_row
            last_row -= 1
        logger.debug(f"No data found in column {column}")
        return 0

    def get_last_non_empty_value_in_range(
        self, start_row: int, column: int, size: int
    ) -> any:
        """
        指定された範囲内で、データが入っている最後の行の値を返します。
        :param start_row: 開始行
        :param column: 列番号
        :param size: 範囲のサイズ（行数）
        :return: データが入っている最後の行の値、全てのセルが空の場合はNone
        """
        try:
            end_row = min(start_row + size - 1, self.worksheet.max_row)
            for row in range(end_row, start_row - 1, -1):
                value = self.worksheet.cell(row=row, column=column).value
                if value is not None and value != "":
                    logger.debug(
                        f"Last non-empty value found in range: row {row}, column {column}, value: {value}"
                    )
                    return value
            logger.debug(
                f"No non-empty data found in the specified range: rows {start_row}-{end_row}, column {column}"
            )
            return None
        except Exception as e:
            logger.error(
                f"An error occurred while getting the last non-empty value in range: {str(e)}"
            )
            return None

    def insert_array_column_wise(self, start_row: int, start_column: int, data: list):
        """
        配列のデータを指定された行から順に列方向に入力します。
        :param start_row: 開始行
        :param start_column: 開始列
        :param data: 入力するデータの配列
        :return: True: 成功、False: エラー発生
        """
        try:
            for i, value in enumerate(data):
                row = start_row + i
                ExcelCellHandler.update_cell(self.worksheet, row, start_column, value)

            logger.debug(
                f"Successfully inserted {len(data)} values column-wise starting from row {start_row}, column {start_column}"
            )
            return True
        except Exception as e:
            logger.error(
                f"An error occurred while inserting data column-wise: {str(e)}"
            )
            return False

    def count_nonempty_cells_in_range(self, column: int, start_row: int, end_row: int):
        """
        指定された列の開始行から終了行までの範囲内の空でないセルの数を数えます。
        :param column: 列番号
        :param start_row: 開始行
        :param end_row: 終了行
        :return: 空でないセルの数
        """
        try:
            count = sum(
                1
                for row in range(start_row, end_row + 1)
                if not pd.isna(self.worksheet.cell(row=row, column=column).value)
                and self.worksheet.cell(row=row, column=column).value != ""
            )
            logger.debug(
                f"Counted {count} non-empty cells in column {column} from row {start_row} to {end_row}"
            )
            return count
        except Exception as e:
            logger.error(f"An error occurred while counting non-empty cells: {str(e)}")
            return 0

    def insert_array_row_wise(self, row: int, start_column: int, data: list):
        """
        配列のデータを指定された行の指定された列から順に行方向に入力します。
        :param row: 入力を開始する行
        :param start_column: 入力を開始する列
        :param data: 入力するデータの配列
        :return: True: 成功、False: エラー発生
        """
        try:
            for i, value in enumerate(data):
                column = start_column + i
                ExcelCellHandler.update_cell(self.worksheet, row, column, value)

            logger.debug(
                f"Successfully inserted {len(data)} values row-wise starting from row {row}, column {start_column}"
            )
            return True
        except Exception as e:
            logger.error(f"An error occurred while inserting data row-wise: {str(e)}")
            return False

    def insert_array_vertically(self, start_row: int, column: int, data: list):
        """
        配列のデータを指定された行の指定された列から順に列方向（縦方向）に入力します。
        :param start_row: 入力を開始する行
        :param column: 入力する列
        :param data: 入力するデータの配列
        :return: True: 成功、False: エラー発生
        """
        try:
            for i, value in enumerate(data):
                row = start_row + i
                ExcelCellHandler.update_cell(self.worksheet, row, column, value)

            logger.debug(
                f"Successfully inserted {len(data)} values vertically starting from row {start_row}, column {column}"
            )
            return True
        except Exception as e:
            logger.error(f"An error occurred while inserting data vertically: {str(e)}")
            return False
